#!/usr/bin/env python
from openpyxl import load_workbook
from openpyxl import Workbook
import glob
import os,sys
from openeye.oechem import *
import psycopg2
import psycopg2.extras
import traceback
from parse import parse



if __name__ == "__main__":
    if len(sys.argv)!=3:
        print "Usage:%s input.xls output.xls"%sys.argv[0]
        print "First row contains data tag n" \
              "ames"
        sys.exit(0)

    OEThrow.SetLevel(OEErrorLevel_Error)
    xlsFile = sys.argv[1]
    xlsFileOut = sys.argv[2]

    wb = load_workbook(xlsFile)
    solid_sheet = wb.get_active_sheet()
    colNames = []
    first_row = solid_sheet.rows[0]
    for c in first_row:
        colName = c.value
        colNames.append(colName.encode('utf-8').strip())

    rowDict = {}
    selectedBatch = []
    maxLotDict = {}
    batch_col_id = colNames.index("BATCH_ID")
    amound_col_id = colNames.index("AMOUNT")
    bionumberList = []
    for row in solid_sheet.rows[1:]:
        batch_id = row[batch_col_id].value
        amount = float(row[amound_col_id].value)
        if amount < 150:
            continue

        p = parse("BIO-{bionumber}-{lot}",batch_id)
        bionumber = "BIO-%s"%p["bionumber"]
        lot = int(p["lot"])
        if bionumber not in bionumberList:
            bionumberList.append(bionumber)

        rowDict[batch_id] = row
        if maxLotDict.has_key(bionumber):
            if maxLotDict[bionumber] < lot:
                maxLotDict[bionumber] = lot
        else:
            maxLotDict[bionumber] = lot

    new_wb = Workbook()
    ws = new_wb.active
    ws.title = "Unique"
    for id,col in enumerate(colNames[0:6]):
        ws.cell(row=1,column=id+1,value=colNames[id])

    for id,bionumber in enumerate(bionumberList):
        lot = maxLotDict[bionumber]
        key = "%s-%02d"%(bionumber,lot)
        row = rowDict[key]
        for col in range(0,6):
            ws.cell(row=id+2,column=col+1,value=row[col].internal_value)
    new_wb.save(xlsFileOut)
