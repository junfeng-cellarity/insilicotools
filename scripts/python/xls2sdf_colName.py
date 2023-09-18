#!/usr/bin/env python
from openpyxl import load_workbook
import glob
import os,sys
from openeye.oechem import *
import psycopg2
import psycopg2.extras
import traceback
from parse import parse

class BiogenDb:
    def __init__(self):
        self.conn = psycopg2.connect(database='biogen2',user='medchem',host='javelin',password='medchem')

    def getMolFromBio(self, bio_number):
        cursor = None
        try:
            cursor = self.conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cursor.execute("select m from rdk.mols where bio_number = %s",(bio_number,))
            result = cursor.fetchone()
            if result is not None:
                smiles = result['m']
                mol = OEGraphMol()
                ifs = oemolistream()
                ifs.SetFormat(OEFormat_SMI)
                ifs.openstring(smiles)
                OEReadMolecule(ifs,mol)
                return mol

            else:
                return None
        except:
            cursor.execute("rollback")
            traceback.print_exc()
            return None
        finally:
            if cursor is not None:
                cursor.close()

    def __del__(self):
        self.conn.close()



if __name__ == "__main__":
    if len(sys.argv)!=4:
        print "Usage:%s input.xls output.sdf colNameWithBionumber"%sys.argv[0]
        print "First row contains data tag names"
        print "BIO-NUMBER is contained in colName specified"
        sys.exit(0)

    OEThrow.SetLevel(OEErrorLevel_Error)
    xlsFile = sys.argv[1]
    sdfFile = sys.argv[2]
    selectedColName = sys.argv[3]

    ofs = oemolostream()
    ofs.open(sdfFile)
    dict = {}
    wb = load_workbook(xlsFile, read_only = True, data_only = True)
    active_sheet = wb.get_active_sheet()
    colNames = []
    row_list = list(active_sheet.iter_rows())
    first_row = row_list[0]
    for c in first_row:
        colName = c.value
        if colName is not None:
            colNames.append(colName.encode('utf-8').strip())
    bio_number_col_id = colNames.index(selectedColName)
    if bio_number_col_id == -1:
        print "Can't find column"
        print colNames,
        print selectedColName
        sys.exit(1)
    db = BiogenDb()
    for row in row_list[1:]:
        s = str(row[bio_number_col_id].value)
        p = parse("BIO-{bionumber}-{lot}",s)
        if p is None:
            p = parse("BIO-{bionumber}",s)
        if p is None:
            print s," is empty"
            continue
        bio_number = "BIO-%s"%p['bionumber']
        # if dict.has_key(bio_number):
        #     continue
        # else:
        #     dict[bio_number] = 1
        mol = db.getMolFromBio(bio_number)
        if mol is None:
            print bio_number, " failed to find."
            continue
        mol.SetTitle(bio_number)
        OESetSDData(mol,"BIO-NUMBER", bio_number)

        for colId,colName in enumerate(colNames):
            if colId != bio_number_col_id:
                value = row[colId].value
                if type(value) == unicode:
                    value = value.encode("utf-8")
                OESetSDData(mol, colNames[colId], str(value))
        if mol is not None:
            OEWriteMolecule(ofs,mol)

    ofs.close()

