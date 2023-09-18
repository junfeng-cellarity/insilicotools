#!/usr/bin/env python
from openeye.oechem import *
import sys,glob,os
import psycopg2
import psycopg2.extras
from openpyxl import load_workbook

DIRECTORY = "/Users/jfeng1/TTBK1"

if __name__ == "__main__":
    if len(sys.argv)!=3:
        print "Usage:%s plate_id.txt bio_number.txt"%(sys.argv[0])
    else:
        plate_ids_tmp = open(sys.argv[1],"r").read().splitlines()
        bio_numbers = open(sys.argv[2],"r").read().splitlines()
        plate_ids = []
        for t in plate_ids_tmp:
            plate_ids.append(t.split()[0])
        print plate_ids,bio_numbers

        selected_bionumbers = []

        plateDb = {}
        files = glob.glob(os.path.join(DIRECTORY,"Screening Set Copy 1 Compilation.xlsx"))
        for f in files:
            if os.path.basename(f).startswith("~"):
                continue
            wb = load_workbook(f)
            progress = 0
            for sheet_name in wb.get_sheet_names():
                active_sheet = wb.get_sheet_by_name(sheet_name)
                for rowid,row in enumerate(active_sheet.rows):
                    plateName = None
                    bio_number = None
                    progress += 1
                    sys.stderr.write("\r%d molecules processed."%progress)
                    sys.stderr.flush()

                    if sheet_name == "Result":
                        if rowid != 0:
                            plateName = str(row[0].value)
                            bio_number = row[4].value

                    else:
                        plateName = str(row[1].value)
                        bio_number = row[7].value

                    if bio_number is not None:
                        bio_number = bio_number.rsplit('-', 1)[0]
                        if plateName is not None and plateName in plate_ids:
                            if bio_number not in selected_bionumbers:
                                selected_bionumbers.append(bio_number)
            break

        files = glob.glob(os.path.join(DIRECTORY,"fssi_platemap.xlsx"))
        for f in files:
            if os.path.basename(f).startswith("~"):
                continue
            wb = load_workbook(f)
            progress = 0
            for sheet_name in wb.get_sheet_names():
                if sheet_name != "Sheet1":
                    continue

                active_sheet = wb.get_sheet_by_name(sheet_name)
                for rowid,row in enumerate(active_sheet.rows):
                    plateName = None
                    bio_number = None
                    progress += 1
                    sys.stderr.write("\r%d molecules processed."%progress)
                    sys.stderr.flush()

                    if rowid != 0:
                        plateName = str(row[12].value)
                        bio_number = row[2].value

                    if bio_number is not None:
                        bio_number = bio_number.rsplit('-', 1)[0]
                        if plateName is not None and plateName in plate_ids:
                            if bio_number not in selected_bionumbers:
                                selected_bionumbers.append(bio_number)
            break

        print len(selected_bionumbers), " are selected."
        print "Already selected:"
        for b in bio_numbers:
            if b in selected_bionumbers:
                print b