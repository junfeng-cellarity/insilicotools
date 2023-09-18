#!/usr/bin/env python
from openpyxl import *
import glob
import os,sys
from openeye.oechem import *
import psycopg2
import psycopg2.extras
import traceback
from parse import parse

if __name__ == "__main__":
    if len(sys.argv)!=3:
        print "Usage:%s inventory.csv excel_to_split.xlsx"
        sys.exit(1)

    csvFile = sys.argv[1]
    xlsFile = sys.argv[2]

    inventory_db = {}
    lines = open(csvFile,"r").read().splitlines()

    for lineno,line in enumerate(lines):
        if lineno == 0:
            continue
        args = line.split(",")
        corporate_id = args[0].replace("\"","")
        inventory_db[corporate_id] = 1

    wb = load_workbook(xlsFile)

    in_stock_wb = Workbook()
    out_stock_wb = Workbook()

    in_stock_ws = in_stock_wb.active
    out_stock_ws = out_stock_wb.active

    solid_sheet = wb.get_active_sheet()
    colNames = []
    first_row = solid_sheet.rows[0]
    for c in first_row:
        if c.value is not None:
            colName = c.value
            colNames.append(colName.encode('utf-8').strip())

    batch_id = colNames.index("VIAL_BARCODE")
    in_stock_ws.append(colNames)
    out_stock_ws.append(colNames)
    n = 0
    for row in solid_sheet.rows[1:]:
        batch = str(row[batch_id].value)
        if inventory_db.has_key(batch):
            in_stock_ws.append(row)
            print  "%s found."%batch
            n+=1
        else:
            out_stock_ws.append(row)
            print  "%s not found."%batch

    in_stock_wb.save("instock.xlsx")
    out_stock_wb.save("outstock.xlsx")




