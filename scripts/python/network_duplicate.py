"""import the package to manipulate the cells directly"""

import openpyxl
wb = openpyxl.load_workbook('node_interactions3.xlsx')
sheet = wb.get_sheet_by_name('node_interactions')

"""Compound Count"""
wb = openpyxl.load_workbook('MOA_for_ErrB_Her_pathway.xlsm')
sheet2 = wb.get_sheet_by_name('MOA_for_ErrB_Her_pathway.txt')
#wb = openpyxl.load_workbook('Bioactives_set.xlsx')
#sheet2 = wb.get_sheet_by_name('Full Bioactive set')

"""Duplicate Bio Numbers"""
#wb = openpyxl.load_workbook('MOA_for_ErrB_Her_pathway.xlsm')
#sheet3 = wb.get_sheet_by_name('MOA_for_ErrB_Her_pathway.txt')

from Tkinter import *
top = Tk()
frame=Frame(top,width=300,height=300)
frame.grid(row=0,column=0)
canvas=Canvas(frame,bg='#FFFFFF',width=1000,height=1000,scrollregion=(0,0,2000, 2000))
hbar=Scrollbar(frame,orient=HORIZONTAL)
hbar.pack(side=BOTTOM,fill=X)
hbar.config(command=canvas.xview)
vbar=Scrollbar(frame,orient=VERTICAL)
vbar.pack(side=RIGHT,fill=Y)
vbar.config(command=canvas.yview)
canvas.config(width=1000,height=1000) #need small canvas to show horizontal bar
canvas.config(xscrollcommand=hbar.set, yscrollcommand=vbar.set)
canvas.pack(side=LEFT,expand=True,fill=BOTH)
#canvas = Canvas(width=800, height= 800, bg='white')
#canvas.pack(expand=YES, fill=BOTH)

"""make all the cell values uppercase"""
for cell in sheet.columns[0]:
    cell.value = cell.value.upper()

for cell in sheet.columns[1]:
    cell.value = cell.value.upper()

for cell in sheet2.columns[0]:
    cell.value = cell.value.upper()

"""dictionary with the bio number as the key and the number of occurences as the value"""
dupBioNumber = {}

for cell in sheet2.columns[0]:
    if cell.value not in dupBioNumber:
        dupBioNumber[cell.value] = 1
    else:
        dupBioNumber[cell.value] += 1
#print("Bio Number duplicates: ", dupBioNumber)
#print()

"""list with each item being a combination of the bio number + '-'  + number of occurences. Ex: 'BIO-0556746-1'"""
dup = []
for bio in dupBioNumber.keys():
    bio = bio + "-" + str(dupBioNumber[bio])
    dup.append(bio)
#print("Now the Bio Number duplicates(keys): ", dup)
#print()

"""target list is a dictionary with the key being the target and the value being the number of occurences aka compound count"""
targetList = {}

"""bioNumber is a dictionary with the key being the target and the value being all the bio numbers and its number of occurences associated with that target"""
bioNumber = {}
i = 1

"""replaces all the bio numbers with the bio number and number of occurences combo"""
for cell in sheet2.columns[0]:
    for bio in dup:
        if (cell.value == bio[:-2]):
            cell.value = bio
            
"""combines all the bionumbers associated with a ttarget together as 1 value"""                         
for cell in sheet2.columns[2]:
    if (cell.value not in targetList) and (cell.value not in bioNumber):
        targetList[cell.value.upper()] = 1
        bioNumber[cell.value.upper()] = str(sheet2['A'+str(i)].value) + ", "
        #bioNumber[cell.value] = sheet2['D'+str(i)].value + ", "
        i += 1
    else:
        targetList[cell.value.upper()] += 1
        temp = str(sheet2['A'+str(i)].value) + ", "
        bioNumber[cell.value.upper()] += temp
        #temp = sheet2['D'+str(i)].value + ", "
        #bioNumber[cell.value] += temp
        i += 1
#print("Compund Count: ", targetList)
#print()
#print ("Length of Compound Count: ", len(targetList))
#print()
#print("bioNumber is: ", bioNumber)
#print()

"""iterator and temporary variables"""
i = 0
child = ""
existingNodes = {}

"""initialize the lists that contain column 1 and column2's contents"""
targets1 = []
targets2 = []

start_x = 400
#start_y = 200
start_y = 20

"""put all the values from the first column into targets1"""
for cell in sheet.columns[0]:
    targets1.append(str(cell.value).upper())

"""put all the values from the second column into targets2"""
for cell in sheet.columns[1]:
    targets2.append(str(cell.value).upper())

"""remove node1 and node2"""
targets1.remove('NODE1')
targets2.remove('NODE2')


for n in targets1:
    for m in bioNumber.keys():
        if (n == m):
            print("Bio Numbers for ", n, "is ", bioNumber.get(n))
            #print(bioNumber[m])
            #return bioNumber[m]


class Node():
    
    def __init__(self, target):
        #self.bioNumber = bioNumber
        self.target = target
        self.parent = None
        self.children = []
        self.x = start_x
        self.y = start_y
        self.visible = False

        
    """bioNumberFunc takes the target as an input and checks to see if it matches with any of the keys in bioNumber and if it does,
        it prints the value - the bio number + occurences
    """
    def bioNumberFunc(self):
        for m in bioNumber.keys():
            print m,self.target
            if (self.target == m):
                print(bioNumber.get(m))


    def numOfChildren(self):
        print("Total number of Nodes are: ", len(self.children))

    def calculateX(self,idx,total,x1,x2):
        div = (x2-x1)/total
        #print (div)
        return x1+idx*div+div/2

    def addChild(self, childNode): # recalculate child x coordinates, when new child added
        self.children.append(childNode)
        childNode.parent = self
        if childNode.y==start_y:
            childNode.y = 100+self.y
        total_span = 100*len(self.children)+500
        x1 = (self.x - total_span/2)
        x2 = (self.x + total_span/2)
        for id,child in enumerate(self.children):
            if child.x==start_x:
                child.x = self.calculateX(id,len(self.children),x1,x2)

    def displayNode(self):
        buttons = {}
        if self.visible:
            return
        self.visible = True
        x1 = self.x-35
        y1 = self.y-25
        x2 = self.x+35
        y2 = self.y+25
        canvas.create_oval(x1,y1+10,x2,y2+10,fill="light blue")
        b = Button(canvas, text = self.target)
        buttons[b] = self.target
        #bioNumber[b] = self.target
        def myfunction(event):
            target =  buttons[event.widget].split(":")[0]
            if bioNumber.has_key(target):
                print bioNumber[target]

        b.bind("<Button-1>", myfunction)
        b.place = (self.x, self.y)
        canvas.create_window(self.x,self.y,anchor=NW,window=b)
        #canvas.create_text(self.x,self.y, text = self.target)
        for child in self.children:
            canvas.create_line(self.x,self.y,child.x,child.y-15,fill="black",arrow=LAST)
            child.displayNode()
            #print(len(self.children))

"""Updates the targets to include the Compound Count from the MOA dataset"""
for n in targets1:
    for m in targetList.keys():
        if (n == m):
            indexx = targets1.index(n)
            targets1[indexx] = m + ": " + str(targetList[m])
#print("targets1: ", targets1)

for n in targets2:
    for m in targetList.keys():
        if (n == m):
            indexx = targets2.index(n)
            targets2[indexx] = m + ": " + str(targetList[m])
#print("targets2: ", targets2)

target_list = []
for target in targets1:
    if target not in target_list:
        target_list.append(target)
        existingNodes[target] = Node(target)

for target in targets2:
    if target not in target_list:
        target_list.append(target)
        existingNodes[target] = Node(target)

for id,target in enumerate(targets1):
   existingNodes[target].addChild(existingNodes[targets2[id]])

existingNodes[targets1[0]].displayNode()
top.mainloop()
