#!/usr/bin/env python
from openpyxl import load_workbook
import xlrd
import glob,os
import sqlite3
from sqlitedict import SqliteDict
import csv
CSV = True
XLS = False
if __name__ == "__main__":
    tareweight_db = SqliteDict("/Users/jfeng1/BiogenDB/VialsWeight/TareWeight2.db",autocommit=True)
    if CSV:
        files = glob.glob(os.path.join("/Users/jfeng1/BiogenDB/VialsWeight/batch9","*.csv"))
        for f in files:
            with open(f,"rb") as csvfile:
                csv_reader = csv.DictReader(csvfile)
                for row in csv_reader:
                    tareweight_db[row['Barcode']]=row['Weight']
#    tareweight_db = SqliteDict("/Volumes/DbUCdd/databases/ADME/TareWeightDB/TareWeight.db",autocommit=True)
    if XLS:
        files = glob.glob(os.path.join("/Users/jfeng1/BiogenDB/VialsWeight/batch8","*.xls"))
        count = 0
        for f in files:
            book = xlrd.open_workbook(f)
            sheet = book.sheet_by_index(0)
            print f,sheet.name,sheet.nrows,sheet.ncols
            # ## for new format
            # for i in range(0,sheet.nrows):
            #     key =  sheet.cell(i,0).value
            #     value = sheet.cell(i,1).value
            #     if type(key) is unicode and type(value) is float:
            #         key = key.encode("utf-8")
            #         tareweight_db[key] = value
            #         print key,value
            # ## for old format ##
            for i in range(28,279):
                try:
                    cell_value = sheet.cell(i, 6).value
                    key = cell_value
                    value = str(sheet.cell(i,10).value)
                    if len(key.strip())==0:
                        continue
                    print count,key,value
                    count += 1
                    tareweight_db[key]=value
                except:
                    import traceback
                    traceback.print_exc()
    # print len(tareweight_db)
    print tareweight_db["BN00151393"] #5.8930
    print tareweight_db["BN00184170"]
    tareweight_db.close()
    # print tareweight_db["BN00079861"] #5.8882
    # print tareweight_db['BN00076551'] #5.9024
    # print tareweight_db['BN00131181'] #5.8846
    # print tareweight_db['BN00141203'] #5.8783



