#!/usr/bin/env python
from openpyxl import load_workbook
import glob
import os,sys
from openeye.oechem import *
import psycopg2
import psycopg2.extras
import traceback
from parse import parse
import re

missingDict = {
    "BIO-0004232":"FC1=CNC(=O)NC1=O",
    "BIO-0504061":"Clc1ccc(cc1)-c2nn(c3ncnc(c32)N)C(C)(C)C",
    "BIO-0499470":"O(C)c1c(ccc(c1)C=CC(=O)CC(=O)C=Cc2cc(c(cc2)O)OC)O",
    "BIO-0552473":"s1c(nc2c1CC(CC2)NCCC)N",
    "BIO-0552576":"n1(nccn1)-c2cc(ccc2)Nc3nc(ncc3C(=O)N)N[C@@H]4[C@@H](CCCC4)N",
    "BIO-0919019":"[Mn+3]3([O-]c1c(cccc1OC)C=NCCN=Cc2c(c(ccc2)OC)[O-]3)[Cl-]"
}

class BiogenDb:
    def __init__(self):
        self.conn = psycopg2.connect(database='biogen2',user='medchem',host='javelin',password='medchem')

    def getMolFromBio(self, bio_number):
        cursor = None
        try:
            cursor = self.conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cursor.execute("select m from rdk.mols where bio_number = %s",(bio_number,))
            result = cursor.fetchone()
            if result is not None:
                smiles = result['m']
                mol = OEGraphMol()
                ifs = oemolistream()
                ifs.SetFormat(OEFormat_SMI)
                ifs.openstring(smiles)
                OEReadMolecule(ifs,mol)
                return mol

            else:
                if missingDict.has_key(bio_number):
                    smiles = missingDict[bio_number]
                    mol = OEGraphMol()
                    ifs = oemolistream()
                    ifs.SetFormat(OEFormat_SMI)
                    ifs.openstring(smiles)
                    OEReadMolecule(ifs,mol)
                    return mol

                return None
        except:
            cursor.execute("rollback")
            traceback.print_exc()
            return None
        finally:
            if cursor is not None:
                cursor.close()

    def __del__(self):
        self.conn.close()

def getWellLocation(well):
    if well is not None:
        try:
            p = re.compile('([A-Za-z]+)([0-9]+)')
            m = p.match(well)
            if m is not None:
                return m.group(1),m.group(2)
        except:
            return None
    return None


working_dir = "/Users/jfeng1/forPatrick"
plateMap_Current = "20151215 Bioactives and Kinase Screening Plates.xlsx"
plateMap_All = "20151124 Bioactives Selleck.xlsx"

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print "Usage: %s raw.csv"%sys.argv[0]
        sys.exit(1)

    plateValuesAll = sys.argv[1]

    db = BiogenDb()

    compoundDict = {}
    xlsFile = os.path.join(working_dir,plateMap_Current)
    wb = load_workbook(xlsFile)
    active_sheet = wb.get_active_sheet()
    colNames = []
    first_row = active_sheet.rows[0]
    for c in first_row:
        colName = c.value
        if colName is not None:
            colNames.append(str(colName))

    for row in active_sheet.rows[1:]:
        colId = colNames.index("BATCH_ID")
        s = str(row[colId].value)
        p = parse("BIO-{bionumber}-{lot}",s)
        bio_number = "BIO-%s"%p['bionumber']
        wellId = colNames.index("POSITION")
        plateId = colNames.index("PLATE/RACK BARCODE")
        compoundDict["%s_%s"%(str(row[plateId].value),str(row[wellId].value))] = bio_number

    print len(compoundDict)

    xlsFile = os.path.join(working_dir,plateMap_All)
    wb = load_workbook(xlsFile)
    active_sheet = wb.get_active_sheet()
    colNames = []
    first_row = active_sheet.rows[0]
    for c in first_row:
        colName = c.value
        if colName is not None:
            colNames.append(str(colName))

    for row in active_sheet.rows[1:]:
        colId = colNames.index("BATCH_ID")
        s = str(row[colId].value)
        p = parse("BIO-{bionumber}-{lot}",s)
        bio_number = "BIO-%s"%p['bionumber']
        print bio_number
        wellId = colNames.index("POSITION")
        plateId = colNames.index("PLATE/RACK BARCODE")
        compoundDict["%s_%s"%(str(row[plateId].value),str(row[wellId].value))] = bio_number

    ofs = oemolostream()
    ofs.open("output.sdf")
    #f = os.path.join(working_dir,plateValuesAll)
    f = plateValuesAll

    rawData = open(f, "r").read().splitlines()

    for line in rawData:
        if line.startswith("Plate"):
            continue
        p = parse("{Plate},{Barcode},{Well},{CalcResult},{Signal},{FlashesTime},{MeasTime},,,,,,,,,,,,,,,,,",line)
        if p is not None and p['Barcode'] and p['Well'] is not None:
            barcode = p['Barcode']
            plateId = re.sub(r"[A|B]$","",barcode)
            wellId = p['Well']
            if wellId == "0":
                continue
            row,col=getWellLocation(wellId)
            key = "%s_%s"%(plateId,wellId)
            if compoundDict.has_key(key):
                bio_number = compoundDict[key]
                mol = db.getMolFromBio(bio_number)
                if mol is None:
                    print "%s not found in db."%bio_number
                    continue
                mol.SetTitle(bio_number)
                OESetSDData(mol, "Well",wellId)
                OESetSDData(mol, "Plate", plateId)
                OESetSDData(mol, "Barcode", barcode)
                OESetSDData(mol, "Plate Row", row)
                OESetSDData(mol, "Plate Col", col)
                OESetSDData(mol, "BIO-NUMBER",bio_number)
                OESetSDData(mol, "Smiles",OEMolToSmiles(mol))
                OESetSDData(mol, "WELL_ROLE","Compound")

                if mol is None:
                    print bio_number," not found in database."
                    continue
                OESetSDData(mol, "CalcResult",p['CalcResult'])

                OEWriteMolecule(ofs,mol)
            else:
                if col=="01" or col=="23":
                    mol = OEGraphMol()
                    OESmilesToMol(mol,"CCCC(=O)[O-].[Na+]")
                    mol.SetTitle("Sodium Butyrate")
                    OESetSDData(mol, "Well",wellId)
                    OESetSDData(mol, "Plate", plateId)
                    OESetSDData(mol, "Barcode", barcode)
                    OESetSDData(mol, "Plate Row", row)
                    OESetSDData(mol, "Plate Col", col)
                    OESetSDData(mol, "CalcResult",p['CalcResult'])
                    OESetSDData(mol, "BIO-NUMBER","N/A")
                    OESetSDData(mol, "Smiles",OEMolToSmiles(mol))
                    OESetSDData(mol, "WELL_ROLE","Positive Control")
                    OEWriteMolecule(ofs,mol)
                elif col =="02" or col=="24":
                    mol = OEGraphMol()
                    OESmilesToMol(mol,"CS(=O)C")
                    mol.SetTitle("DMSO")
                    OESetSDData(mol, "Well",wellId)
                    OESetSDData(mol, "Plate", plateId)
                    OESetSDData(mol, "Barcode", barcode)
                    OESetSDData(mol, "Plate Row", row)
                    OESetSDData(mol, "Plate Col", col)
                    OESetSDData(mol, "CalcResult",p['CalcResult'])
                    OESetSDData(mol, "BIO-NUMBER","N/A")
                    OESetSDData(mol, "Smiles","CS(=O)C")
                    OESetSDData(mol, "WELL_ROLE","Negative Control")
                    OEWriteMolecule(ofs,mol)

    ofs.close()



