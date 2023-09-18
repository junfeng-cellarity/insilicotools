#!/usr/bin/env python
import sys,os
import openpyxl
from openpyxl import load_workbook
if __name__ == "__main__":
    if len(sys.argv)!=4:
        print "Usage:%s input.xlsx tags.txt output.xlsx"%sys.argv[0]
    else:
        xlsFile = sys.argv[1]
        tagsTxtFile = sys.argv[2]
        outputXlsFile = sys.argv[3]

        keys = open(tagsTxtFile,"r").read().splitlines()
        print keys
        colNameDict = {}
        newWb = openpyxl.Workbook()
        wb = load_workbook(xlsFile)
        new_sheet = newWb.create_sheet(0)
        sheet = wb.get_active_sheet()
        for rowid,row in enumerate(sheet.rows):
            if rowid ==0:
                new_sheet.append(row)
            else:
                for colId,col in enumerate(row):
                    if colId == 0:
                        if col.value in keys:
                            new_sheet.append(row)
                            print "Found:", col.value
        newWb.save(sys.argv[3])







